import openpyxl
from SeedTrak.settings import BASE_DIR
from main.models import Semilla


# from main.scripts import save_semilla_from_xl
def save_semilla_from_xl():
    path = '{}{}'.format(BASE_DIR, '/main/scripts/files/semillas.xlsx')
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    for i in range(1, sheet.max_row + 1):
        codigo = sheet.cell(row=i, column=1).value
        nombre = sheet.cell(row=i, column=2).value
        Semilla.objects.get_or_create(codigo=codigo, nombre=nombre)
    